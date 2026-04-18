import io
from datetime import datetime
from typing import List

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database.models import Listing

BOOL_MAP = {True: "Да", False: "Нет", None: "—"}


async def build_excel(listings: List[Listing]) -> io.BytesIO:
    rows = []
    for l in listings:
        rows.append(
            {
                "Источник": l.source,
                "Район": l.district or "—",
                "Адрес": l.address or "—",
                "Цена (€/мес)": l.price,
                "Площадь (м²)": l.area,
                "Планировка": l.room_structure or "—",
                "Вода включена": BOOL_MAP[l.water_included],
                "Цена воды (€)": l.water_price or "—",
                "Электричество включено": BOOL_MAP[l.electricity_included],
                "Доступно с": l.available_from or "—",
                "Арендодатель": l.lessor_name or "—",
                "Частник": BOOL_MAP[l.is_private_lessor],
                "Ссылка": l.url,
                "Дата парсинга": (
                    l.scraped_at.strftime("%Y-%m-%d %H:%M") if l.scraped_at else "—"
                ),
            }
        )

    df = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Listings")
        ws = writer.sheets["Listings"]

        # Шапка — синий фон, белый жирный текст
        header_fill = PatternFill("solid", fgColor="2E5F8A")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Чередующиеся строки
        light_fill = PatternFill("solid", fgColor="EBF2FA")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = light_fill

        # Ссылки кликабельные
        url_col = df.columns.get_loc("Ссылка") + 1
        for row_idx, listing in enumerate(listings, start=2):
            cell = ws.cell(row=row_idx, column=url_col)
            cell.hyperlink = listing.url
            cell.value = "Открыть"
            cell.font = Font(color="0563C1", underline="single")

        # Ширина колонок по содержимому
        for col_idx, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_len + 4, 40
            )

    output.seek(0)
    return output
